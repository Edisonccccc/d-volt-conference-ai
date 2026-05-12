"""Conference attendee scoring — turn an xlsx of names into a ranked list.

Workflow:
1. ``parse_attendee_xlsx`` reads a workbook, finds the header row, maps the
   common columns (First Name / Last Name / Company), and returns a list
   of normalized row dicts. Extra columns are preserved in `raw_row`.
2. ``score_attendee`` runs a Claude+web_search agent per attendee:
     - web search the company (for fit signals vs. the seller's offering)
     - web search the person (for role + seniority)
     - call ``submit_attendee_score`` with score 1-4 + reason + briefs.
   The seller-company context blob (cached from earlier work) is fed in
   so the scoring is tailored to the rep's actual go-to-market.
3. ``score_conference_attendees`` is the batch orchestrator: a small
   ThreadPoolExecutor that runs N candidates in parallel and writes the
   results back to the attendees table as each one finishes. Designed to
   be called from a FastAPI BackgroundTask so the request returns
   immediately and the rep polls for progress.
"""

from __future__ import annotations

import io
import logging
import os
from concurrent.futures import ThreadPoolExecutor
from typing import Any, Dict, List, Optional, Tuple

from . import storage
from .company_profile import get_or_fetch_company_profile
from .costing import anthropic_cost
from .llm import make_client, stream_to_terminal
from .rate_limit import attendee_bucket


log = logging.getLogger("conference-ai.attendee-scoring")
MODEL = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-5")

# Concurrency cap. The token bucket protects against rate-limit blowups,
# so we can keep parallelism reasonably high — bucket will throttle when
# needed. Default 5; tune via ATTENDEE_SCORING_PARALLELISM env var.
SCORING_PARALLELISM = int(os.getenv("ATTENDEE_SCORING_PARALLELISM", "5"))

# Per-call token estimates used to reserve budget BEFORE the call.
# Real numbers from observed runs: ~5K input (prompt + web search results)
# and ~600 output. We bake in some headroom on input so the reservation
# isn't too optimistic — actual ITPM consumption is the bigger constraint.
EST_INPUT_TOKENS  = int(os.getenv("ATTENDEE_EST_INPUT_TOKENS",  "6000"))
EST_OUTPUT_TOKENS = int(os.getenv("ATTENDEE_EST_OUTPUT_TOKENS", "1000"))


# ---------------------------------------------------------------------------
# xlsx parser
# ---------------------------------------------------------------------------

# Column-header aliases — what the user might have written maps to our
# canonical key. Match is case-insensitive after stripping whitespace.
_HEADER_ALIASES = {
    "first_name": {"first name", "firstname", "first", "given name", "given"},
    "last_name":  {"last name", "lastname", "last", "surname", "family name", "family"},
    "company":    {"company", "organization", "organisation", "employer", "firm", "company name"},
}


def _normalize_header(s: Any) -> str:
    return (str(s or "")).strip().lower()


def parse_attendee_xlsx(blob: bytes) -> List[Dict[str, Any]]:
    """Parse the bytes of an .xlsx upload into normalized attendee rows.

    Looks for the header row in the first 5 rows of the active sheet,
    detects which columns are First Name / Last Name / Company by
    fuzzy-matching headers, and yields a dict per row with the canonical
    keys plus all unmapped extras grouped under their original headers.

    Rows missing both first_name and last_name AND company are dropped.
    """
    # Imported here so the module loads even if openpyxl isn't installed —
    # only matters when this function is actually called.
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # Find the header row in the first 5 lines. The header row is the one
    # where at least one cell matches one of our known column aliases.
    rows_buffer: List[Tuple] = []
    header_idx = -1
    header_map: Dict[int, str] = {}   # column index -> canonical key
    header_raw: Dict[int, str] = {}   # column index -> original header text
    for idx, row in enumerate(rows_iter):
        rows_buffer.append(row)
        cells = [_normalize_header(c) for c in row]
        # Heuristic: a row is "header-like" if any cell matches an alias.
        candidate_map: Dict[int, str] = {}
        candidate_raw: Dict[int, str] = {}
        for col_i, cell in enumerate(cells):
            for canon, aliases in _HEADER_ALIASES.items():
                if cell in aliases:
                    candidate_map[col_i] = canon
                    candidate_raw[col_i] = str(row[col_i] or "").strip()
                    break
            if cell and col_i not in candidate_map:
                # Keep the original header text for "extras" passthrough.
                candidate_raw[col_i] = str(row[col_i] or "").strip()
        if len(candidate_map) >= 2:
            header_idx = idx
            header_map = candidate_map
            header_raw = candidate_raw
            break
        if idx >= 4:   # only peek the first 5 rows
            break

    if header_idx < 0:
        log.warning("parse_attendee_xlsx: no recognizable header row found; "
                    "falling back to positional (col 0=first, 1=last, 2=company).")
        # Positional fallback for headerless sheets — assume the user's
        # ordering matches what we expect (per the sample).
        header_idx = -1
        header_map = {0: "first_name", 1: "last_name", 2: "company"}
        header_raw = {0: "First Name", 1: "Last Name", 2: "Company"}

    out: List[Dict[str, Any]] = []
    # If there was a real header, body starts at header_idx+1; else
    # everything we've already buffered counts as data.
    body_start = header_idx + 1 if header_idx >= 0 else 0

    def emit(row_tuple: Tuple) -> None:
        rec: Dict[str, Any] = {}
        extras: Dict[str, Any] = {}
        for col_i, val in enumerate(row_tuple):
            text = "" if val is None else str(val).strip()
            if not text:
                continue
            canon = header_map.get(col_i)
            if canon:
                rec[canon] = text
            else:
                key = header_raw.get(col_i) or f"col_{col_i}"
                extras[key] = text
        # Drop empty rows.
        if not (rec.get("first_name") or rec.get("last_name") or rec.get("company")):
            return
        if extras:
            rec.update(extras)
        out.append(rec)

    # Emit the buffered rows past the header.
    for i in range(body_start, len(rows_buffer)):
        emit(rows_buffer[i])
    # Drain the remaining rows from the iterator.
    for row in rows_iter:
        emit(row)
    return out


# ---------------------------------------------------------------------------
# AI scoring
# ---------------------------------------------------------------------------

SYSTEM = (
    "You triage sales-conference attendees for a B2B seller. The goal is "
    "to find REAL DEMAND signals: companies that look like they need what "
    "the seller offers, especially right now. You use web search to study "
    "the COMPANY — its products, recent news (last 90 days), expansion or "
    "M&A activity, hiring trends, technical priorities — and score the "
    "candidate 1-4. Role of the specific rep is secondary; their company's "
    "demand signal is what matters. You are biased toward 'skip' or 'don't "
    "meet' when evidence is thin — conferences are short and reps' time "
    "is finite. Never invent biographical facts about a real individual."
)


def _ws_tool(max_uses: int = 3) -> dict:
    return {
        "type": "web_search_20250305",
        "name": "web_search",
        "max_uses": max_uses,
    }


SUBMIT_SCORE = {
    "name": "submit_attendee_score",
    "description": (
        "Deliver the verdict for one conference attendee. Call exactly once "
        "after gathering enough info from web search."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "score": {
                "type": "integer",
                "enum": [1, 2, 3, 4],
                "description": (
                    "Priority on the seller's meeting shortlist. "
                    "4 = MUST meet (clear ICP fit AND decision-maker). "
                    "3 = GOOD to meet (good company fit, role uncertain or junior). "
                    "2 = CAN SKIP (peripheral fit; only if time allows). "
                    "1 = DO NOT MEET (no fit, competitor, or irrelevant)."
                ),
            },
            "score_reason": {
                "type": "string",
                "description": (
                    "1-2 sentence factual justification. Cite the concrete "
                    "signal: company fit, role/seniority, market overlap. "
                    "Be terse — this displays next to the name in a list."
                ),
            },
            "company_brief": {
                "type": "string",
                "description": (
                    "One-sentence factual blurb about the company: what "
                    "they do and who they sell to."
                ),
            },
            "rep_brief": {
                "type": "string",
                "description": (
                    "1-2 sentence factual blurb about this specific person: "
                    "role and responsibility. Empty string if you can't "
                    "find them confidently — DO NOT invent."
                ),
            },
            "sources": {
                "type": "array",
                "items": {"type": "string"},
                "description": "URLs you actually consulted.",
            },
        },
        "required": ["score", "score_reason"],
    },
}


def _attendee_prompt(
    first_name: Optional[str],
    last_name: Optional[str],
    company: Optional[str],
    seller_company: str,
    seller_context: str,
) -> str:
    full_name = " ".join(filter(None, [first_name, last_name])) or "(unknown name)"
    return f"""\
SELLER (your client): {seller_company}
SELLER CONTEXT (what they sell, who they target):
{seller_context or "(no context available)"}

ATTENDEE:
- Name:    {full_name}
- Company: {company or "(unknown)"}

GOAL: Score the COMPANY for the seller's conference meeting shortlist.
Focus on whether {seller_company}'s offering matches a real, current need.

Steps:
1. Search the company. Pull recent signals (last ~90 days strongly
   preferred): news, press releases, executive announcements, M&A,
   technology partnerships, RFPs, expansion, hiring spikes, regulatory
   filings — anything that hints at active demand for what
   {seller_company} sells.
2. Cross-reference what the company does (products / services /
   customer base) against the seller's offering. Is there a plausible
   buyer-seller fit? Is this an existing or near-term need, or
   speculative?
3. DO NOT run a separate dedicated search for the individual rep. If
   the rep's name + title appears organically in the company-side
   results, you may use it; otherwise leave `rep_brief` empty. Time
   spent on rep lookups is better spent on demand signals.
4. Score 1-4 weighted heavily on company-side demand signal; rep role
   is a tie-breaker, not the primary driver.

Scoring guide:
- 4 (MUST):       clear ICP fit AND a concrete recent signal indicating
                  active demand (a real opening to walk into).
- 3 (GOOD):       solid ICP fit, no specific timing signal — still worth
                  the conversation.
- 2 (CAN SKIP):   peripheral / speculative fit, no urgency.
- 1 (DON'T MEET): clearly no fit (e.g. competitor, irrelevant industry,
                  wrong scale).

Default toward lower scores when signals are thin. Cite the SPECIFIC
signal you found in `score_reason` ("Closed Series C in Feb; expanding
manufacturing capacity" beats "looks like a fit"). Call
`submit_attendee_score` exactly once.
"""


def score_attendee(
    *,
    first_name: Optional[str],
    last_name: Optional[str],
    company: Optional[str],
    seller_company: str,
    seller_context: str,
) -> Tuple[dict, float]:
    """Run one scoring call. Returns ``(result_dict, cost_usd)``.

    Acquires a token-bucket reservation BEFORE the SDK call so concurrent
    workers don't burst past the Anthropic ITPM/OTPM caps. Raises on API
    failure so the caller can mark the row as 'error'.
    """
    client = make_client(timeout=90.0)
    label = f"score-{(last_name or 'x').lower()[:8]}"

    # Proactive rate-limit gate. Blocks until both ITPM + OTPM windows
    # have room for our estimate; the SDK's own retry-on-429 is a backstop.
    bucket = attendee_bucket()
    reservation = bucket.acquire(
        EST_INPUT_TOKENS, EST_OUTPUT_TOKENS, label=label,
    )

    final = None
    try:
        final = stream_to_terminal(
            client,
            label=label,
            model=MODEL,
            max_tokens=1024,
            system=SYSTEM,
            tools=[_ws_tool(3), SUBMIT_SCORE],
            tool_choice={"type": "auto"},
            messages=[{
                "role": "user",
                "content": _attendee_prompt(
                    first_name, last_name, company,
                    seller_company, seller_context,
                ),
            }],
        )
    finally:
        # True up the reservation with actuals (or zero if the call
        # never produced usage — keeps the bucket from staying
        # over-reserved when the SDK raises before we get a response).
        usage = getattr(final, "usage", None) if final is not None else None
        actual_in  = int(getattr(usage, "input_tokens",  0) or 0)
        actual_out = int(getattr(usage, "output_tokens", 0) or 0)
        # web_search adds input tokens via tool results; include cache
        # accounting too so the bucket doesn't under-count.
        actual_in += int(getattr(usage, "cache_read_input_tokens",     0) or 0)
        actual_in += int(getattr(usage, "cache_creation_input_tokens", 0) or 0)
        bucket.record(reservation, actual_in, actual_out)

    cost = anthropic_cost(getattr(final, "usage", None), MODEL)
    for block in final.content:
        if (getattr(block, "type", None) == "tool_use"
            and getattr(block, "name", None) == "submit_attendee_score"):
            return dict(block.input or {}), cost

    # The call succeeded but the model didn't emit the expected tool_use
    # block. Treat as soft failure with a manual-review fallback.
    log.warning("score_attendee: model didn't call submit_attendee_score "
                "for %r %r — recording manual-review fallback.",
                first_name, last_name)
    return {
        "score": 2,
        "score_reason": "Could not produce a scored verdict; manually review.",
    }, cost


def score_conference_attendees(
    conf_id: str, seller_company: Optional[str],
) -> None:
    """Batch entrypoint — score every pending attendee on a conference.

    Loads the seller-company context once, then fans out the per-attendee
    scoring calls. Designed for FastAPI BackgroundTask; the request that
    triggers this returns ``202 Accepted`` and the rep polls the
    conference detail endpoint for updated rows.
    """
    seller = (seller_company or "the seller").strip() or "the seller"
    # Reuse the cached seller-profile blob from earlier work.
    context_blob = get_or_fetch_company_profile(seller) or ""

    pending = storage.list_pending_attendees(conf_id)
    if not pending:
        return

    log.info(
        "attendee-scoring: starting %d candidates for conference %s "
        "(seller=%r, parallelism=%d)",
        len(pending), conf_id, seller, SCORING_PARALLELISM,
    )

    def _work(att) -> None:
        try:
            storage.update_attendee_status(att.id, "researching")
            result, cost = score_attendee(
                first_name=att.first_name,
                last_name=att.last_name,
                company=att.company,
                seller_company=seller,
                seller_context=context_blob,
            )
            score = result.get("score")
            try:
                score_int = int(score) if score is not None else None
            except (TypeError, ValueError):
                score_int = None
            if score_int is None:
                # Successful API call but no usable score — treat as error so
                # the row stays visible and retryable rather than silently
                # landing in a "scored with NULL" black hole.
                raise ValueError("scoring call returned no usable score")
            if not (1 <= score_int <= 4):
                score_int = max(1, min(4, score_int))
            storage.update_attendee_score(
                att.id,
                score=score_int,
                score_reason=(result.get("score_reason") or None),
                company_brief=(result.get("company_brief") or None),
                rep_brief=(result.get("rep_brief") or None),
                sources=list(result.get("sources") or []),
                cost_usd=cost,
            )
        except Exception as exc:  # noqa: BLE001
            log.exception("attendee-scoring: %s failed: %s", att.id, exc)
            storage.update_attendee_status(
                att.id, "error", error=f"{exc.__class__.__name__}: {exc}",
            )

    with ThreadPoolExecutor(max_workers=SCORING_PARALLELISM) as ex:
        list(ex.map(_work, pending))

    log.info("attendee-scoring: conference %s done.", conf_id)
